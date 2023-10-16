"""
Parses the CubeSat requirements that were provided by the 
University of Phoenix.
"""

from typing import Dict, List
import re
import openpyxl

class Requirement:
    """
    Data class to hold information about a requirement
    """
    def __init__(self, rid, title, text, parents) -> None:
        self.id = rid
        self.title = title
        self.text = text
        self.parents = parents

class RequirementsParser:
    """
    Parses the requirements spreadsheet and
    outputs a graph in DOT notation.
    """
    
    def __init__(self) -> None:
        # This is where the spreadsheet is located
        # (relative to the root folder)
        self.filename = 'static/phoenix_requirements.xlsx'
    
    def _is_requirements_id(self, req_id:str) -> bool:
        """
        Checks if the provided value looks like a requirment.
        This helps filtering out rows in the spreadsheet that don't
        contain requirements (e.g. headers).
        """
        if req_id is None:
            return False
        else:
            return bool(re.search(r'\d', req_id))
     
     
    def _read_spreadsheet(self) -> List[Requirement]:
        """
        Parses the input spreadsheet and returns
        a list of Requirement objects.
        """
        reqs:List[Requirement] = []
        
        # Open spreadsheet using OPENPYXL
        wb = openpyxl.load_workbook(self.filename)
        sheet = wb[wb.sheetnames[0]]
        for row in sheet.iter_rows(2):
            
            # Read the values from each cell
            req_id = row[0].value
            title = row[1].value
            text = row[2].value
            parents = []
            if row[4].value is not None:
                parents = row[4].value.split('\n')
            
            # Only add row if this looks like a requirement
            if self._is_requirements_id(req_id) and title is not None and text is not None:
                reqs.append(Requirement(req_id, title, text, parents))
        # Return all requirements
        return reqs

    
    def _print_graph(self, requirements:List[Requirement]) -> None:
        """
        Prints a graph that illustrates the dependencies
        between requirements.
        """
        print('digraph Phoenix_Cubesat {')
        for req in requirements:
            for parent in req.parents:
                print(f'"{parent}" -> "{req.id}"')
        print("}")
        
               
    def parse(self):
        """
        Main function for parsing requirements.
        """
        requirements = self._read_spreadsheet()
        self._print_graph(requirements)
        
        
if __name__ == '__main__':
    RequirementsParser().parse()